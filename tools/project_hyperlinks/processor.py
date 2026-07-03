import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

ISSUE_URL_TEMPLATE = "https://tracker.rddm.team/issues/{number}"
HYPERLINK_FONT = Font(color="0563C1", underline="single")

SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}


def ensure_supported_extension(input_path: str) -> None:
    ext = Path(input_path).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Неподдерживаемое расширение: {ext}\n"
            f"Поддерживаются: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )


def normalize_issue_number(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text or None


def display_value(number: str):
    return int(number) if number.isdigit() else number


def output_path(input_path: Path, output_dir: Path | None = None) -> Path:
    directory = output_dir if output_dir is not None else input_path.parent
    candidate = directory / f"Hyper_{input_path.stem}.xlsx"
    if not candidate.exists():
        return candidate
    return directory / f"Hyper_{input_path.stem}1.xlsx"


def issue_url(number: str) -> str:
    return ISSUE_URL_TEMPLATE.format(number=number)


def set_issue_hyperlink(cell, number: str) -> None:
    url = issue_url(number)
    cell.value = display_value(number)
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url, display=number)
    cell.font = HYPERLINK_FONT


def apply_hyperlinks_to_sheet(sheet) -> None:
    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_index, column=1)
        number = normalize_issue_number(cell.value)
        if number is None:
            continue
        set_issue_hyperlink(cell, number)


def read_csv_rows(input_path: Path) -> list[list[str]]:
    with input_path.open("r", encoding="utf-8-sig", newline="") as src:
        sample = src.read(4096)
        src.seek(0)
        delimiter = ","
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            delimiter = dialect.delimiter
        except csv.Error:
            pass
        return list(csv.reader(src, delimiter=delimiter))


def csv_to_excel(input_path: Path, output_file: Path) -> None:
    rows = read_csv_rows(input_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Issues"

    for row in rows:
        sheet.append(row)

    apply_hyperlinks_to_sheet(sheet)
    workbook.save(output_file)


def process_excel(input_path: Path, output_file: Path) -> None:
    workbook = load_workbook(input_path)
    for sheet in workbook.worksheets:
        apply_hyperlinks_to_sheet(sheet)
    workbook.save(output_file)


def process_file(input_path: str, output_dir: Path | None = None) -> Path:
    ensure_supported_extension(input_path)
    source = Path(input_path)
    destination = output_path(source, output_dir=output_dir)

    if source.suffix.lower() == ".csv":
        csv_to_excel(source, destination)
    else:
        process_excel(source, destination)

    return destination
